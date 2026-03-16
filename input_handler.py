"""
Read book inputs from Excel and upsert to Supabase.
Reads an Excel file (.xlsx) with columns: title, notes_on_outline_before
Creates or updates book records in the database.
"""

import os
from openpyxl import Workbook, load_workbook
import db

def read_input(filepath: str) -> list[dict]:
    """
    Read book inputs from an Excel file and create records in Supabase.

    Expected Excel columns:
      - title (required)
      - notes_on_outline_before (required — notes guiding outline generation)

    Returns a list of created/updated book records.
    """
    if not os.path.exists(filepath):
        print(f"[ERROR] Input file not found: {filepath}")
        return []

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    # Read header row to find column indexes
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_lower = [h.lower().strip() if h else "" for h in headers]

    try:
        title_idx = headers_lower.index("title")
    except ValueError:
        print("[ERROR] Excel file must have a 'title' column.")
        return []

    try:
        notes_idx = headers_lower.index("notes_on_outline_before")
    except ValueError:
        print("[ERROR] Excel file must have a 'notes_on_outline_before' column.")
        return []

    books = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[title_idx]
        notes = row[notes_idx] if notes_idx < len(row) else None

        if not title:
            continue  # Skip empty rows

        title = str(title).strip()
        notes = str(notes).strip() if notes else None

        # Create book in Supabase
        book = db.create_book(title, notes)
        books.append(book)
        print(f"  [OK] Created book: '{title}' (ID: {book['id']})")

    wb.close()
    print(f"\n[DONE] Loaded {len(books)} book(s) from {filepath}")
    return books


def create_sample_input(filepath: str) -> None:
    """Create a sample input Excel file for demonstration."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Books Input"

    # Headers
    ws.append(["title", "notes_on_outline_before"])

    # Sample data
    ws.append([
        "The Future of Artificial Intelligence",
        "Focus on practical applications in healthcare, education, and business. "
        "Include historical context of AI development. Make it accessible to "
        "non-technical readers. Target 5-6 chapters. Include real-world case studies."
    ])

    wb.save(filepath)
    print(f"[OK] Sample input file created: {filepath}")